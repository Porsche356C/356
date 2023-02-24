from openpyxl import load_workbook
from yattag import Doc, indent
import os

menufile = "xml/menu_vw.xml"
# Load our Excel File
wb = load_workbook("config/vw_parts.xlsx")

catalog = {
	"1" : "Engine",
	"2" : "Transmission",
	"3" : "Chassis",
	"4" : "Fuel Tank & Wheels",
	"5" : "Body",
	"6" : "Bumpers & Doors",
	"7" : "Seats & Trim",
	"8" : "Soft & Hard Tops",
	"9" : "Electrical Equipment"
}


def add_menu_group(i, mdoc, mtag, mtext):
	sheet = wb[wb.sheetnames[i]]
	file = sheet.title.split("-")
	group = file[0]
	groupCurrent = group[5:]
	print(f"i: {i}  group: {catalog[groupCurrent]}")

	with mtag('group'):	
		mdoc.attr(id = groupCurrent)
		mdoc.attr(title = catalog[groupCurrent])
		i = add_menu_section(i, mdoc, mtag, mtext)
	return i


def add_menu_section(i, mdoc, mtag, mtext):
	sheet = wb[wb.sheetnames[i]]
	file = sheet.title.split("-")
	group = file[0]
	groupCurrent = group[5:]
	groupPrevious = groupCurrent

	sectionnum = file[1]
	section = "section" + sectionnum
	filename = sheet.title
	path = "catalog/" + group + "/" + section + "/"
	title = sheet.cell(row=2, column=1).value
	print(f"i: {i}  title: {title}")
	
	while ( groupCurrent == groupPrevious ) and ( i <= len(wb.sheetnames )):
		with mtag('section', ('data-dir', path)):
			mdoc.attr(id = sectionnum)
			mdoc.attr(title = title)
			# mdoc.attr("data-dir=" , mtext(path))
			print(f"path: {path}")

			with mtag("diagram"): mtext("vw_parts.png")
			with mtag("areaFile"): mtext("vw_overlay.html")
			with mtag("partslist"): mtext("vw_parts.xml")
		i += 1
		if i < len(wb.sheetnames):
			sheet = wb[wb.sheetnames[i]]
			file = sheet.title.split("-")
			group = file[0]
			groupCurrent = group[5:]

			sectionnum = file[1]
			section = "section" + sectionnum
			filename = sheet.title
			path = "catalog/" + group + "/" + section + "/"
			title = sheet.cell(row=2, column=1).value
			print(f"i: {i}  title: {title}")
		else:
			break
	return i


def build_menu():
	menu = open(menufile, "w").close()
	menu = open(menufile, "a")
	groupPrevious = ""
	groupCurrent = ""

	# create instance of yattag for menu file
	mdoc, mtag, mtext = Doc().tagtext()

	with mtag('parts-catalog'):
		i = 0
		while i < len(wb.sheetnames):
			i = add_menu_group(i, mdoc, mtag, mtext)
			print(f"i: {i}")

			
	result = indent(
		mdoc.getvalue(),
		indentation='\t',
		indent_text=False
	)

	print(result)

	menu.write(result)
	menu.close()

	
def get_part(row, doc, tag, text):
	with tag("part") :
		with tag("diagram_name"): text(row[0])
		with tag("ref"): text(row[1])
		with tag("website") : text(row[2])
		with tag("group"): text(row[3])
		with tag("item"): text(row[4])
		with tag("description"): text(row[5])
		with tag("partno"): text(row[6])
		with tag("qty"): text(row[7])
		with tag("image"): text("images/",row[8])
		with tag("note"): text(row[9])

def write_partslists():
	for sheet in wb:
		# create instance of yattag for the catalog partslists
		doc, tag, text = Doc().tagtext()

		file = sheet.title.split("-")
		group = file[0]
		section = "section" + file[1]
		filename = "vw_parts.xml"
		path = "catalog/" + group + "/" + section + "/"

		with tag('partslist'):
			for row in sheet.iter_rows(min_row=2):
				row = [cell.value for cell in row]
				get_part(row, doc, tag, text)

		result = indent(
			doc.getvalue(),
			indentation='\t',
			indent_text=False
		)

		if not os.path.exists(path):
			os.makedirs(path)

		with open(path + filename, "w") as f:
			f.write(result)
			f.close()




if __name__ == "__main__":
	build_menu()
	write_partslists()

	
