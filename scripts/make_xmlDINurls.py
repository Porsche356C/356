from openpyxl import load_workbook
from yattag import Doc, indent

# Load our Excel File
wb = load_workbook("xml/din_urls.xlsx")

# xml_header = '<?xml version="1.0" encoding="UTF-8"?>'
# xml_ns = '<din:specs xmlns:din="http://356.miplace.com/specs">'
# xml_closetag = '</din:specs>'

# for sheet in wb:
sheet = wb.worksheets[0]
# Returning returns a triplet
doc, tag, text = Doc().tagtext()

# Appends the String to document
# doc.asis(xml_header)
# doc.asis(xml_ns)

with tag('dins'):
	for row in sheet.iter_rows(min_row=2):
		row = [cell.value for cell in row]
		with tag("din_id") :
			doc.attr(id = row[0])
			doc.attr(image = "images/"+row[2])
			doc.attr(url = row[3])
			# with tag("id"): text(row[0])
			# with tag("description"): text(row[1])
			# with tag("image") : text("images/", row[2])
			# with tag("url"): text(row[3])
# doc.asis(xml_closetag)

result = indent(
	doc.getvalue(),
	indentation='\t',
	indent_text=False
)

with open("xml/din_specs.xml", "w") as f:
	f.write(result)
	f.close()
	

